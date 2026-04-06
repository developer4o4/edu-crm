"""
Hisobotlar va Excel export
"""
import io
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_style(wb):
    """Excel stil yordamchisi"""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1D9E75', end_color='1D9E75', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    return header_font, header_fill, header_align, thin_border


class PaymentReportView(APIView):
    """To'lovlar hisoboti - JSON va Excel"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.payments.models import Payment
        from django.db.models import Sum

        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        export = request.query_params.get('export', 'json')

        qs = Payment.objects.filter(status=Payment.Status.PAID).select_related(
            'student', 'group', 'received_by'
        )

        if date_from:
            qs = qs.filter(paid_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(paid_at__date__lte=date_to)

        qs = qs.order_by('-paid_at')

        if export == 'excel':
            return self._export_excel(qs)

        data = [{
            'id': str(p.id),
            'student': p.student.full_name,
            'phone': p.student.phone,
            'group': p.group.name if p.group else '—',
            'amount': float(p.amount),
            'payment_type': p.get_payment_type_display(),
            'month': p.month.strftime('%B %Y') if p.month else '—',
            'paid_at': p.paid_at.strftime('%d.%m.%Y %H:%M') if p.paid_at else '—',
            'received_by': p.received_by.get_full_name() if p.received_by else '—',
        } for p in qs]

        total = qs.aggregate(total=Sum('amount'))['total'] or 0

        return Response({
            'count': len(data),
            'total_amount': float(total),
            'payments': data,
        })

    def _export_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "To'lovlar"
        ws.sheet_view.rightToLeft = False

        header_font, header_fill, header_align, thin_border = create_excel_style(wb)

        headers = [
            "T/r", "O'quvchi", "Telefon", "Guruh",
            "Summa (UZS)", "To'lov turi", "Oy", "To'langan sana", "Qabul qildi"
        ]
        col_widths = [5, 25, 16, 18, 16, 14, 14, 20, 20]

        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = "TO'LOVLAR HISOBOTI"
        title_cell.font = Font(bold=True, size=14, color='1D9E75')
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:I2')
        ws['A2'] = f"Yaratilgan: {timezone.now().strftime('%d.%m.%Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='right')
        ws.row_dimensions[2].height = 18

        # Headers row 4
        for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = width
        ws.row_dimensions[4].height = 22

        # Data
        total_amount = 0
        for row_num, payment in enumerate(queryset, 5):
            row_data = [
                row_num - 4,
                payment.student.full_name,
                payment.student.phone,
                payment.group.name if payment.group else '—',
                float(payment.amount),
                payment.get_payment_type_display(),
                payment.month.strftime('%B %Y') if payment.month else '—',
                payment.paid_at.strftime('%d.%m.%Y %H:%M') if payment.paid_at else '—',
                payment.received_by.get_full_name() if payment.received_by else '—',
            ]
            total_amount += float(payment.amount)

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if col_num == 5:  # Amount column
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')

        # Total row
        total_row = len(queryset) + 5
        ws.cell(row=total_row, column=1, value="JAMI:")
        ws.merge_cells(f'A{total_row}:D{total_row}')
        total_cell = ws.cell(row=total_row, column=5, value=total_amount)
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
        total_cell.fill = PatternFill(start_color='E1F5EE', end_color='E1F5EE', fill_type='solid')

        # Freeze header
        ws.freeze_panes = 'A5'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"payments_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class AttendanceReportView(APIView):
    """Davomat hisoboti"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.attendance.models import Attendance, AttendanceSession
        from apps.groups.models import Group

        group_id = request.query_params.get('group_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        export = request.query_params.get('export', 'json')

        sessions = AttendanceSession.objects.all()
        if group_id:
            sessions = sessions.filter(group_id=group_id)
        if date_from:
            sessions = sessions.filter(date__gte=date_from)
        if date_to:
            sessions = sessions.filter(date__lte=date_to)

        sessions = sessions.prefetch_related('records__student').order_by('-date')

        if export == 'excel':
            return self._export_excel(sessions)

        data = []
        for session in sessions:
            records = session.records.all()
            data.append({
                'session_id': session.id,
                'group': session.group.name,
                'date': session.date.strftime('%d.%m.%Y'),
                'topic': session.topic,
                'total': records.count(),
                'present': records.filter(status='present').count(),
                'absent': records.filter(status='absent').count(),
                'late': records.filter(status='late').count(),
            })

        return Response({'sessions': data})

    def _export_excel(self, sessions):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Davomat"
        header_font, header_fill, header_align, thin_border = create_excel_style(wb)

        headers = ["Sana", "Guruh", "Mavzu", "Jami", "Keldi", "Kelmadi", "Kech keldi", "%"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row_num, session in enumerate(sessions, 2):
            records = session.records.all()
            total = records.count()
            present = records.filter(status='present').count()
            absent = records.filter(status='absent').count()
            late = records.filter(status='late').count()
            rate = round((present / total * 100) if total else 0, 1)

            row = [
                session.date.strftime('%d.%m.%Y'),
                session.group.name,
                session.topic or '—',
                total, present, absent, late, rate
            ]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        return response


class MonthlyIncomeReportView(APIView):
    """Oylik/yillik daromad hisoboti"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.payments.models import Payment
        from django.db.models import Sum
        from django.db.models.functions import TruncMonth

        year = int(request.query_params.get('year', timezone.now().year))
        export = request.query_params.get('export', 'json')

        monthly = Payment.objects.filter(
            status=Payment.Status.PAID,
            paid_at__year=year,
        ).annotate(paid_month=TruncMonth('paid_at')).values('paid_month').annotate(
            total=Sum('amount'),
            count=__import__('django.db.models', fromlist=['Count']).Count('id'),
        ).order_by('paid_month')

        yearly_total = Payment.objects.filter(
            status=Payment.Status.PAID,
            paid_at__year=year,
        ).aggregate(total=Sum('amount'))['total'] or 0

        data = [
            {
                'month': item['paid_month'].strftime('%B %Y'),
                'month_num': item['paid_month'].month,
                'total': float(item['total'] or 0),
                'count': item['count'],
            }
            for item in monthly
        ]

        if export == 'excel':
            return self._export_excel(data, year, float(yearly_total))

        return Response({
            'year': year,
            'yearly_total': float(yearly_total),
            'monthly': data,
        })

    def _export_excel(self, data, year, yearly_total):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{year} yil daromad"
        header_font, header_fill, header_align, thin_border = create_excel_style(wb)

        ws['A1'] = f"{year} YIL DAROMAD HISOBOTI"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        headers = ["Oy", "Jami (UZS)", "Tranzaksiyalar soni", "O'rtacha (UZS)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = 22

        for row_num, item in enumerate(data, 4):
            avg = item['total'] / item['count'] if item['count'] else 0
            row = [item['month'], item['total'], item['count'], round(avg)]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                if col in (2, 4):
                    cell.number_format = '#,##0'

        # Total
        total_row = len(data) + 4
        ws.cell(row=total_row, column=1, value="YILLIK JAMI").font = Font(bold=True)
        tc = ws.cell(row=total_row, column=2, value=yearly_total)
        tc.font = Font(bold=True)
        tc.number_format = '#,##0'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="income_{year}.xlsx"'
        return response