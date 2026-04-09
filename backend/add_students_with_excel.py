"""
O'quvchilarni Excel dan import qilish skripti

ISHLATISH:
  1. Shu faylni backend/ papkasiga qo'ying
  2. Excel faylini ham backend/ papkasiga qo'ying
  3. Pastdagi o'zgaruvchilarni to'ldiring
  4. docker-compose exec backend python import_students.py

EXCEL FORMATI:
  A - Ism
  B - Familiya
  C - Telefon raqam
  D, E, F... - Davomat (ixtiyoriy, import qilinmaydi)
"""

import os, django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from apps.students.models import Student, User
from apps.groups.models import Group, GroupMembership

# ================================================================
# === SHUNING O'ZINI O'ZGARTIRING ================================
# ================================================================

EXCEL_FILE  = 'students.xlsx'   # Excel fayl nomi (backend/ papkasida)
SHEET_NAME  = None              # None = birinchi sheet, yoki 'Sheet1' deb yozing
GROUP_NAME  = 'Python 4'        # Guruh nomi (DB da mavjud bo'lishi kerak)
ADDRESS     = 'Paxtaobod tuman' # Manzil
SKIP_ROWS   = 1                 # Nechta qatorni o'tkazib yuborish (sarlavha)

# ================================================================

def normalize_phone(raw):
    """998901234567 → +998901234567"""
    if not raw:
        return None
    phone = str(raw).strip().replace(' ', '').replace('-', '')
    # float bo'lib kelishi mumkin: 998901234567.0
    if phone.endswith('.0'):
        phone = phone[:-2]
    if phone.startswith('+'):
        return phone
    if phone.startswith('998'):
        return '+' + phone
    if phone.startswith('8') and len(phone) == 10:
        return '+99' + phone
    if len(phone) == 9:
        return '+998' + phone
    return '+' + phone

def create_user_for_student(student):
    username = student.phone.replace('+', '')
    password = username[-4:]
    if User.objects.filter(username=username).exists():
        user = User.objects.get(username=username)
        if not student.user:
            student.user = user
            student.save()
        return user, password
    user = User.objects.create_user(
        username=username, password=password,
        first_name=student.first_name, last_name=student.last_name,
        role='student',
    )
    student.user = user
    student.save()
    return user, password

def main():
    # Excel faylini ochish
    if not os.path.exists(EXCEL_FILE):
        print(f"XATO: '{EXCEL_FILE}' fayli topilmadi!")
        print("Excel faylini backend/ papkasiga ko'chiring")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
    print(f"Sheet: {ws.title}, Jami qator: {ws.max_row}")

    # Guruhni topish
    group = None
    if GROUP_NAME:
        try:
            group = Group.objects.get(name=GROUP_NAME, is_active=True)
            print(f"Guruh topildi: {group.name}")
        except Group.DoesNotExist:
            print(f"OGOHLANTIRISH: '{GROUP_NAME}' guruh topilmadi — guruhsiz qo'shiladi")
        except Group.MultipleObjectsReturned:
            group = Group.objects.filter(name=GROUP_NAME, is_active=True).first()
            print(f"Guruh topildi (birinchisi): {group.name}")

    created = 0
    skipped = 0
    errors  = []

    rows = list(ws.iter_rows(min_row=SKIP_ROWS + 1, values_only=True))

    for i, row in enumerate(rows, start=SKIP_ROWS + 1):
        # Bo'sh qatorni o'tkazib yuborish
        if not any(row[:3]):
            continue

        first_name = str(row[0] or '').strip()
        last_name  = str(row[1] or '').strip()
        raw_phone  = row[2]

        if not first_name and not last_name:
            continue

        phone = normalize_phone(raw_phone)
        if not phone:
            errors.append(f"  Qator {i}: Telefon yo'q — {first_name} {last_name}")
            continue

        # Mavjudligini tekshirish
        if Student.objects.filter(phone=phone).exists():
            student = Student.objects.get(phone=phone)
            print(f"  MAVJUD: {student.full_name} ({phone})")
            # Guruhga qo'shish (agar yo'q bo'lsa)
            if group and not GroupMembership.objects.filter(student=student, group=group, is_active=True).exists():
                if not group.is_full:
                    GroupMembership.objects.create(student=student, group=group)
                    print(f"    → {group.name} guruhiga qo'shildi")
            skipped += 1
            continue

        # Yangi o'quvchi yaratish
        try:
            student = Student.objects.create(
                first_name=first_name,
                last_name=last_name,
                phone=phone,
                parent_phone=phone,
                address=ADDRESS,
                status='active',
            )

            # User yaratish
            user, password = create_user_for_student(student)

            # Guruhga qo'shish
            if group:
                if not group.is_full:
                    GroupMembership.objects.create(student=student, group=group)
                    group_info = f"→ {group.name}"
                else:
                    group_info = "→ Guruh to'la!"
            else:
                group_info = ""

            print(f"  OK: {student.full_name} | tel: {phone} | login: {user.username} | parol: {password} {group_info}")
            created += 1

        except Exception as e:
            errors.append(f"  Qator {i} ({first_name} {last_name}): {e}")

    print(f"\n{'='*50}")
    print(f"Yaratildi:  {created} ta o'quvchi")
    print(f"Mavjud edi: {skipped} ta")

    if errors:
        print(f"\nXATOLAR ({len(errors)} ta):")
        for err in errors:
            print(err)

    print(f"\nDB dagi jami o'quvchilar: {Student.objects.count()} ta")

if __name__ == '__main__':
    main()
